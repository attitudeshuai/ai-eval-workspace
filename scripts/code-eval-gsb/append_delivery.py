#!/usr/bin/env python3
"""
向 GSB 交付 Excel（如 【成都】GSB0731.xlsx）的「数据表」追加一行评测记录。

输入为一个 JSON 文件：{ "列名": "值", ... }，列名必须与 Excel 表头完全一致。
- 表头中不存在的 key 直接报错（防止拼写错误静默丢数据）
- 未提供的列留空，并在输出中列出，便于人工核对
- 默认按「Github Repo」列查重，重复时需 --force 才允许追加
- 「提交时间」缺省时自动填入当前时间

用法：
    python3 append_delivery.py --xlsx <交付表.xlsx> --json <记录.json> [--dry-run] [--force]

依赖：openpyxl（请在工作台 .venv 中安装后运行）
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("错误：需要 openpyxl。请先执行：")
    print("  python3 -m venv .venv && .venv/bin/pip install openpyxl")
    print("然后用 .venv/bin/python 运行本脚本。")
    sys.exit(1)


def main():
    ap = argparse.ArgumentParser(description="向 GSB 交付 Excel 追加一行评测记录")
    ap.add_argument("--xlsx", required=True, help="交付 Excel 文件路径")
    ap.add_argument("--json", required=True, help="记录 JSON 文件路径（{列名: 值}）")
    ap.add_argument("--dry-run", action="store_true", help="只校验不写文件")
    ap.add_argument("--force", action="store_true", help="Github Repo 重复时仍强制追加")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    json_path = Path(args.json)
    if not xlsx_path.exists():
        print(f"错误：Excel 文件不存在：{xlsx_path}")
        sys.exit(1)
    if not json_path.exists():
        print(f"错误：JSON 文件不存在：{json_path}")
        sys.exit(1)

    record = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(record, dict):
        print("错误：JSON 顶层必须是对象 {列名: 值}")
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["数据表"] if "数据表" in wb.sheetnames else wb.worksheets[0]

    headers = [c.value for c in ws[1]]
    header_set = set(headers)

    # 1) 未知列名校验
    unknown = [k for k in record if k not in header_set]
    if unknown:
        print("错误：以下列名在 Excel 表头中不存在（疑似拼写错误）：")
        for k in unknown:
            print(f"  - {k}")
        sys.exit(1)

    # 2) 查重（Github Repo）
    repo_col = headers.index("Github Repo") + 1
    repo_val = str(record.get("Github Repo", "")).strip()
    if repo_val:
        for row in ws.iter_rows(min_row=2, min_col=repo_col, max_col=repo_col):
            if str(row[0].value or "").strip() == repo_val:
                print(f"错误：Github Repo 已存在于第 {row[0].row} 行：{repo_val}")
                print("如确认要重复追加，请加 --force。")
                sys.exit(1)

    # 3) 提交时间缺省自动填充
    record.setdefault("提交时间", datetime.now().strftime("%Y-%m-%d %H:%M"))

    # 4) 汇总留空列
    empty = [h for h in headers if h and not str(record.get(h, "")).strip()]
    print(f"将写入 {len(headers) - len(empty)}/{len(headers)} 列，以下 {len(empty)} 列留空：")
    for h in empty:
        print(f"  - {h}")

    if args.dry_run:
        print("\n[dry-run] 校验通过，未写入文件。")
        return

    # 5) 追加行
    row_values = [record.get(h, "") if h else "" for h in headers]
    ws.append(row_values)
    wb.save(xlsx_path)
    print(f"\n已追加到 {xlsx_path.name}「{ws.title}」第 {ws.max_row} 行。")


if __name__ == "__main__":
    main()
