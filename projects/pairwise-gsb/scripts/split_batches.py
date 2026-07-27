"""
将大 Excel 按批次切分，每批 N 行（默认从 config.toml 读取）。

用法:
    python split_batches.py <source.xlsx> <dest_dir> [--batch-size 15]

示例:
    python split_batches.py sessions/pairwise-gsb/0724/2026-07-27/original.xlsx sessions/pairwise-gsb/0724/2026-07-27
"""

import sys
import os
import json
import argparse
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    sys.exit(1)

from _config import (
    get_batch_size,
    get_batch_prefix,
    get_batch_num_width,
    get_input_columns,
    get_batch_input_dir,
    get_batch_output_dir,
    get_batch_images_dir,
    format_batch_input_file,
    format_batch_output_file,
)


def copy_data_validations(src_ws, dst_ws, start_row, end_row):
    """
    把源工作表的数据验证（下拉/多选列表）复制到目标工作表，
    并将行范围重映射到批次内行号（源行 start_row~end_row-1 → 目标行 2~）。

    这样切分后的批次文件仍保留下拉多选 list，标注输出不会退化为纯文字。
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    dvs = getattr(src_ws, "data_validations", None)
    if not dvs or not dvs.dataValidation:
        return 0

    copied = 0
    for dv in dvs.dataValidation:
        new_dv = DataValidation(
            type=dv.type,
            formula1=dv.formula1,
            allow_blank=dv.allow_blank,
            showErrorMessage=dv.showErrorMessage,
        )
        dst_ws.add_data_validation(new_dv)
        added = False
        for rng in dv.sqref.ranges:
            for col in range(rng.min_col, rng.max_col + 1):
                col_letter = get_column_letter(col)
                for src_row in range(max(rng.min_row, start_row), min(rng.max_row, end_row - 1) + 1):
                    dst_row = src_row - start_row + 2
                    new_dv.add(f"{col_letter}{dst_row}")
                    added = True
        if added:
            copied += 1
    return copied


def split_excel(src_path, dest_dir, batch_size=None):
    if batch_size is None:
        batch_size = get_batch_size()

    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    total_rows = ws.max_row - 1  # 去掉表头
    num_batches = (total_rows + batch_size - 1) // batch_size

    # 校验输入列名
    headers = [str(ws.cell(row=1, column=c).value or "") for c in range(1, ws.max_column + 1)]
    required = get_input_columns()
    missing = [c for c in required if c not in headers]
    if missing:
        print(f"⚠️ 警告：源 Excel 缺少 config.toml 中定义的列: {missing}")
        print(f"   现有列: {headers}")

    print(f"源文件: {src_path}")
    print(f"总数据行: {total_rows}, 每批: {batch_size}, 批次数: {num_batches}")
    print("=" * 60)

    for batch_idx in range(num_batches):
        batch_name = f"{get_batch_prefix()}{batch_idx + 1:0{get_batch_num_width()}d}"
        batch_dir = os.path.join(dest_dir, batch_name)

        os.makedirs(os.path.join(batch_dir, get_batch_input_dir()), exist_ok=True)
        os.makedirs(os.path.join(batch_dir, get_batch_images_dir()), exist_ok=True)
        os.makedirs(os.path.join(batch_dir, get_batch_output_dir()), exist_ok=True)

        batch_wb = openpyxl.Workbook()
        batch_ws = batch_wb.active

        # 复制表头
        for col in range(1, ws.max_column + 1):
            batch_ws.cell(row=1, column=col, value=ws.cell(row=1, column=col).value)

        # 复制本批数据行
        start_row = 2 + batch_idx * batch_size
        end_row = min(start_row + batch_size, ws.max_row + 1)
        for src_row in range(start_row, end_row):
            dst_row = src_row - start_row + 2
            for col in range(1, ws.max_column + 1):
                batch_ws.cell(row=dst_row, column=col, value=ws.cell(row=src_row, column=col).value)

        count = end_row - start_row
        row_start = start_row
        row_end = end_row - 1
        input_filename = format_batch_input_file(row_start, row_end)
        output_filename = format_batch_output_file(row_start, row_end)

        # 复制数据验证（下拉多选列表），行号重映射到批次内
        dv_count = copy_data_validations(ws, batch_ws, start_row, end_row)

        batch_xlsx = os.path.join(batch_dir, get_batch_input_dir(), input_filename)
        batch_wb.save(batch_xlsx)
        batch_wb.close()

        # 生成/更新 metadata.json
        try:
            original_rel = os.path.relpath(src_path, batch_dir)
        except ValueError:
            # Windows 跨盘符时 relpath 会失败，退化为绝对路径
            original_rel = os.path.abspath(src_path)
        metadata = {
            "date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
            "batch_id": batch_name,
            "title": f"GSB 标注 - 第{batch_idx + 1}批",
            "category": "图像生成质量评估",
            "created_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
            "annotator": "AI",
            "status": "pending",
            "source": {
                "original_excel": original_rel,
                "row_range": f"行{row_start}~行{row_end}",
            },
            "summary": {
                "total_items": count,
                "items_with_images": 0,
                "items_without_images": 0,
                "breakdown": {
                    "T2I": 0,
                    "image_editing": 0,
                    "reference_generation": 0,
                },
                "gsb_distribution": {
                    "图片1更好": 0,
                    "图片2更好": 0,
                    "无法区分": 0,
                },
            },
            "files": {
                "input": f"{get_batch_input_dir()}/{input_filename}",
                "images": f"{get_batch_images_dir()}/",
                "output": f"{get_batch_output_dir()}/{output_filename}",
                "manifest": f"{get_batch_images_dir()}/manifest.json",
            },
        }
        metadata_path = os.path.join(batch_dir, "metadata.json")
        with open(metadata_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)

        print(f"  ✅ {batch_name}: 原始行 {start_row}~{end_row - 1} → {count} 条, 数据验证 {dv_count} 组")

    wb.close()
    print("=" * 60)
    print(f"✅ 切分完成: {num_batches} 个批次 → {dest_dir}")

    # 列出所有批次供后续使用
    print(f"\n批次列表: ", end="")
    print(", ".join(f"{get_batch_prefix()}{i + 1:0{get_batch_num_width()}d}" for i in range(num_batches)))


def main():
    parser = argparse.ArgumentParser(description="切分 Excel 为批次")
    parser.add_argument("source", help="源 Excel 路径")
    parser.add_argument("dest_dir", help="目标目录")
    parser.add_argument("--batch-size", type=int, default=None, help="每批行数（默认从 config.toml 读取）")
    args = parser.parse_args()

    if not os.path.exists(args.source):
        print(f"❌ 源文件不存在: {args.source}")
        sys.exit(1)

    split_excel(args.source, args.dest_dir, args.batch_size)


if __name__ == "__main__":
    main()
