"""
GSB 标注结果校验脚本
对 annotated.xlsx 进行自动校验，检测常见错误模式。

用法:
    python validate_annotations.py <batch_dir>
    python validate_annotations.py sessions/pairwise-gsb/0724/2026-07-27/batch-01/
"""

import sys
import os
import json
import re
import argparse
from collections import Counter
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _config import (
    get_tags,
    get_gsb_values,
    get_consistency_values,
    get_output_columns,
    get_batch_output_dir,
    find_batch_output_file,
    get_batch_error_file,
    get_max_error_rate,
    get_tag_separator,
    get_undecided_warn_threshold,
)


def parse_tags(tag_str):
    """解析归因标签字符串。

    以 config 的 tag_separator（默认英文逗号）为主分隔符；
    同时兼容旧数据中的顿号分隔。
    """
    if not tag_str or str(tag_str).strip() == "":
        return []
    sep = get_tag_separator()
    s = str(tag_str).replace("、", sep)
    return [t.strip() for t in s.split(sep) if t.strip()]


def normalize(s):
    """标准化字符串用于匹配"""
    if not s:
        return ""
    return str(s).strip().replace(" ", "").replace("　", "")


def reason_mentions_tag(reason, tag):
    """
    检查 reason 中是否体现了标签对应的问题。
    简单实现：检查标签中的关键词是否在 reason 中出现。
    """
    reason_n = normalize(reason)
    tag_n = normalize(tag)

    # 标签关键词映射（尽量覆盖常见表述，用于启发式提示，不作为硬性判定）
    keyword_map = {
        "指令响应--完全未响应": ["未响应", "没有完成", "未完成", "漏", "缺少", "没做"],
        "指令响应--人像相关元素响应差": ["人像", "人脸", "妆容", "姿势", "景别", "性别", "服饰", "表情", "人物", "脸型", "眼镜"],
        "指令响应--物体/图像等具象元素响应差": ["物体", "动植物", "道具", "背景", "图案", "画幅", "比例", "元素", "形态", "不完整", "主体", "二维码", "定位角"],
        "指令响应--风格/色彩等抽象元素响应差": ["风格", "色彩", "滤镜", "色调", "抽象", "颜色"],
        "指令响应--实体数量响应差": ["数量", "个数", "只", "个", "张", "多", "少", "翻倍", "不足"],
        "指令响应--实体位置响应差": ["位置", "方位", "上下", "左右", "中间", "角落", "排版"],
        "世界/专业知识--场景类型错误": ["场景", "类型错误"],
        "世界/专业知识--名人不像（非给定图，仅用于 T2I）": ["名人", "不像"],
        "世界/专业知识--新概念/流行的/专有的新名词响应错误": ["概念", "名词", "响应错误"],
        "一致性--知名IP保持差": ["IP", "形象", "保持", "变了"],
        "一致性--真人-人物保持差": ["真人", "人物", "保持", "人脸", "特征", "变了", "脸型", "面部", "体态"],
        "一致性--非真实人物-人物保持差": ["非真实", "虚拟", "动画", "人物", "保持", "变了"],
        "一致性--风格/色调保持差": ["风格", "色调", "保持", "变了", "变化"],
        "一致性--其他元素保持差": ["背景", "道具", "元素", "保持", "变了", "丢失", "重绘", "改变", "状态栏", "文字"],
        "视觉效果-存在明显的结构崩坏": ["结构", "崩坏", "手指", "肢体", "错位", "粘连", "多指", "少指"],
        "视觉效果-结构无崩坏但违反物理规律": ["物理", "悬浮", "重力", "反射", "运动", "比例"],
        "视觉效果-自然真实-生成效果不自然（假）": ["不自然", "假", "不真实", "真实感", "乱码", "错字", "逻辑", "合理"],
        "视觉效果-自然真实-AI感": ["AI感", "塑料感", "AI"],
        "视觉效果-自然真实-负向信息带入": ["负向", "不适", "负面"],
        "视觉效果-画质差": ["画质", "模糊", "噪点", "马赛克", "清晰度", "低清"],
        "视觉效果-基础美感差": ["美感", "观感", "美观", "不好看", "构图", "裁切", "黑带", "黑边", "倾斜", "正式感"],
        "视觉效果-设计场景中设计感差": ["设计感", "版式", "配色", "字体", "排版", "海报", "Logo"],
        "指令响应--其他内容响应差": ["其他", "文字", "错字", "乱码", "文案"],
    }

    keywords = keyword_map.get(tag, [])
    if not keywords:
        return False

    for kw in keywords:
        if kw in reason_n:
            return True
    return False


def extract_reason_sections(reason):
    """从 reason 中提取四个段落的内容，支持换行和单行两种格式"""
    sections = {
        "instruction": "",
        "consistency": "",
        "visual": "",
        "conclusion": "",
    }
    if not reason:
        return sections

    # 先按换行切分
    lines = [l.strip() for l in reason.split("\n") if l.strip()]

    # 如果只有一行，再按句号切分尝试识别段落
    if len(lines) == 1:
        parts = [p.strip() for p in reason.split("。") if p.strip()]
        lines = [p + "。" for p in parts]

    for line in lines:
        # 注意顺序：先判结论（结论句可能含"视觉"等词），再判各维度段
        if "综合判断" in line or "综上" in line or line.startswith("因此"):
            if not sections["conclusion"]:
                sections["conclusion"] = line
        elif "指令遵循" in line or line.startswith("指令"):
            if not sections["instruction"]:
                sections["instruction"] = line
        elif "一致性" in line:
            if not sections["consistency"]:
                sections["consistency"] = line
        elif "视觉" in line or "视觉效果" in line:
            if not sections["visual"]:
                sections["visual"] = line

    return sections


def has_template_placeholder(reason):
    """检测 reason 是否包含模板占位符"""
    placeholders = ["图1……", "图2……", "[具体对象]", "[具体问题]", "[关键差异]", "[结论]", "……"]
    return any(p in reason for p in placeholders)


def validate_row(row_idx, data, tags):
    """校验单行标注数据，返回 (errors, warnings)。

    errors：客观可判定的硬性错误（字段非法、T2I 一致性、模板复制等），计入错误率。
    warnings：需要人工/AI 复核确认的主观项（标签-reason 语义对应、"其他"类标签使用），
              关键词启发式无法可靠判定，不计入错误率，由 quality-audit 抽检覆盖。
    """
    errors = []
    warnings = []

    g = str(data.get("整体gsb结果", "")).strip()
    i = str(data.get("分维度gsb/指令遵循", "")).strip()
    it = str(data.get("分维度gsb/指令遵循归因标签", "")).strip()
    c = str(data.get("分维度gsb/一致性", "")).strip()
    ct = str(data.get("分维度gsb/一致性归因标签", "")).strip()
    v = str(data.get("分维度gsb/视觉效果", "")).strip()
    vt = str(data.get("分维度gsb/视觉效果归因标签", "")).strip()
    reason = str(data.get("reason", "")).strip()
    evaluator = str(data.get("评测人", "")).strip()

    VALID_GSB = set(get_gsb_values())
    VALID_CONSISTENCY = set(get_consistency_values())
    VALID_INSTRUCTION_TAGS = tags["instruction"]
    VALID_CONSISTENCY_TAGS = tags["consistency"]
    VALID_VISUAL_TAGS = tags["visual"]

    # === 1. 字段值合法性 ===
    if g not in VALID_GSB:
        errors.append(f"[字段] 整体gsb结果='{g}' 不在有效值 {VALID_GSB} 中")
    if i not in VALID_GSB:
        errors.append(f"[字段] 指令遵循='{i}' 不在有效值中")
    if c not in VALID_CONSISTENCY:
        errors.append(f"[字段] 一致性='{c}' 不在有效值中")
    if v not in VALID_GSB:
        errors.append(f"[字段] 视觉效果='{v}' 不在有效值中")

    # 检查非标准值（如 image1、图1、same）
    non_standard = {"image1", "image2", "图1", "图2", "same", "1", "2", "相同"}
    for field_name, value in [("整体gsb结果", g), ("指令遵循", i), ("一致性", c), ("视觉效果", v)]:
        if value and value in non_standard:
            errors.append(f"[字段] {field_name}='{value}' 是非交付标准值，必须使用'图片1更好/图片2更好/无法区分'")

    # === 2. T2I 任务一致性检查 ===
    if c == "不涉及":
        if ct and ct != "nan":
            errors.append(f"[一致性] T2I 任务一致性='不涉及'但归因标签不为空: '{ct}'")
    else:
        # 非 T2I，一致性归因标签应在一致性标签范围内
        tags_list = parse_tags(ct)
        for t in tags_list:
            if t and t not in VALID_CONSISTENCY_TAGS:
                errors.append(f"[标签] 一致性归因标签 '{t}' 不在有效一致性标签中")

    # === 3. 归因标签合法性 ===
    it_tags = parse_tags(it)
    for t in it_tags:
        if t and t not in VALID_INSTRUCTION_TAGS:
            errors.append(f"[标签] 指令遵循归因标签 '{t}' 不在有效标签中")

    vt_tags = parse_tags(vt)
    for t in vt_tags:
        if t and t not in VALID_VISUAL_TAGS:
            errors.append(f"[标签] 视觉效果归因标签 '{t}' 不在有效标签中")

    # === 4. 评测人检查 ===
    if not evaluator or evaluator == "nan":
        errors.append(f"[字段] 评测人为空")

    # === 5. reason 质量检查 ===
    if not reason or reason == "nan":
        errors.append(f"[reason] reason 为空")
    else:
        sections = extract_reason_sections(reason)

        # 必须包含四段结构
        if not sections["instruction"]:
            errors.append(f"[reason] reason 缺少'指令遵循'部分")
        if c != "不涉及" and not sections["consistency"]:
            errors.append(f"[reason] 非T2I任务reason缺少'一致性'部分")
        if not sections["visual"]:
            errors.append(f"[reason] reason 缺少'视觉效果'部分")
        if not sections["conclusion"]:
            errors.append(f"[reason] reason 缺少结论（需含'综合判断'或'因此'）")

        # 禁止空洞描述
        hollow_patterns = ["图1更好", "图2更好", "无法区分", "存在一致性问题", "视觉效果较差", "整体更自然"]
        if reason in hollow_patterns or (len(reason) < 30 and any(p in reason for p in hollow_patterns)):
            errors.append(f"[reason] reason过于空洞，需写具体对象和问题")

        # 模板复制检测
        if has_template_placeholder(reason):
            errors.append(f"[reason] reason 疑似复制模板未修改，包含占位符")

        # 检查结论与整体GSB一致
        conclusion = sections["conclusion"]
        if conclusion:
            if g == "图片1更好" and "图片2更好" in conclusion and "图片1更好" not in conclusion:
                errors.append(f"[矛盾] 整体GSB='图片1更好'但reason结论倾向图片2")
            if g == "图片2更好" and "图片1更好" in conclusion and "图片2更好" not in conclusion:
                errors.append(f"[矛盾] 整体GSB='图片2更好'但reason结论倾向图片1")
            if g == "无法区分" and ("图片1更好" in conclusion or "图片2更好" in conclusion):
                # 可能是描述分维度，但如果明确写了整体结论则报错
                if "综合判断" in conclusion or "综上" in conclusion:
                    errors.append(f"[矛盾] 整体GSB='无法区分'但reason结论有明确选边")

    # === 6. 标签与 reason 对应性检查（启发式提示，语义对应由 quality-audit 复核） ===
    if reason and reason != "nan":
        for t in it_tags:
            if t and not reason_mentions_tag(reason, t):
                warnings.append(f"[标签-reason] 指令遵循标签 '{t}' 在 reason 中未见明显对应描述，请人工确认")
        for t in vt_tags:
            if t and not reason_mentions_tag(reason, t):
                warnings.append(f"[标签-reason] 视觉效果标签 '{t}' 在 reason 中未见明显对应描述，请人工确认")
        if c != "不涉及":
            for t in parse_tags(ct):
                if t and not reason_mentions_tag(reason, t):
                    warnings.append(f"[标签-reason] 一致性标签 '{t}' 在 reason 中未见明显对应描述，请人工确认")

    # === 7. 重复归因检测 ===
    # 简单检测：同一标签出现在多个维度
    all_tags = []
    if it_tags:
        all_tags.extend([("指令遵循", t) for t in it_tags])
    if c != "不涉及":
        ct_tags = parse_tags(ct)
        if ct_tags:
            all_tags.extend([("一致性", t) for t in ct_tags])
    if vt_tags:
        all_tags.extend([("视觉效果", t) for t in vt_tags])

    tag_counter = Counter(t for _, t in all_tags)
    for tag, count in tag_counter.items():
        if count > 1:
            dims = [d for d, t in all_tags if t == tag]
            errors.append(f"[重复归因] 标签 '{tag}' 同时出现在 {dims} 中")

    # === 8. "无法区分"硬选检测 ===
    if i == "无法区分" and len(it_tags) > 0:
        # 如果 reason 中没有说明两图都有问题或问题相近，则报错
        if reason and "两图" not in reason and "都" not in reason and "相近" not in reason and "相同" not in reason:
            errors.append(f"[硬选] 指令遵循='无法区分'但归因为单边，且 reason 未说明两图问题相近")
    if v == "无法区分" and len(vt_tags) > 0:
        if reason and "两图" not in reason and "都" not in reason and "相近" not in reason and "相同" not in reason:
            errors.append(f"[硬选] 视觉效果='无法区分'但归因为单边，且 reason 未说明两图问题相近")
    if c == "无法区分" and len(parse_tags(ct)) > 0:
        if reason and "两图" not in reason and "都" not in reason and "相近" not in reason and "相同" not in reason:
            errors.append(f"[硬选] 一致性='无法区分'但归因为单边，且 reason 未说明两图问题相近")

    # === 9. "其他内容响应差"使用提醒（是否有更具体标签需人工判断） ===
    if it_tags and "指令响应--其他内容响应差" in it_tags:
        warnings.append(f"[标签] 使用了'指令响应--其他内容响应差'，请确认无更具体标签可用")

    return errors, warnings


def validate_file(xlsx_path):
    """校验整个标注文件"""
    try:
        import openpyxl
    except ImportError:
        print("❌ 需要安装 openpyxl: pip install openpyxl")
        return None

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 读取表头，确定列索引
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            headers[str(cell.value).strip()] = col_idx

    # 列名映射
    col_map = {}
    for col_name in get_output_columns():
        col_map[col_name] = headers.get(col_name)

    # 检查必需列是否存在
    missing = [name for name, idx in col_map.items() if idx is None]
    if missing:
        print(f"⚠️ 警告：Excel 中缺少以下列: {missing}")
        print(f"   现有列: {list(headers.keys())}")

    tags = get_tags()

    all_errors = {}
    all_warnings = {}
    total_rows = 0
    gsb_dist = Counter()

    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for field, col in col_map.items():
            if col is None:
                row_data[field] = ""
                continue
            cell_val = ws.cell(row=row_idx, column=col).value
            row_data[field] = str(cell_val) if cell_val is not None else ""

        # 跳过空行
        if not row_data.get("整体gsb结果") or row_data["整体gsb结果"] == "None":
            continue

        total_rows += 1
        gsb_dist[row_data["整体gsb结果"]] += 1

        errors, warnings = validate_row(row_idx, row_data, tags)
        if errors:
            all_errors[row_idx] = errors
        if warnings:
            all_warnings[row_idx] = warnings

    return {
        "total_rows": total_rows,
        "error_rows": len(all_errors),
        "warning_rows": len(all_warnings),
        "error_rate": len(all_errors) / total_rows if total_rows > 0 else 0,
        "gsb_distribution": dict(gsb_dist),
        "errors": all_errors,
        "warnings": all_warnings,
    }


def main():
    parser = argparse.ArgumentParser(description="GSB 标注结果校验")
    parser.add_argument("batch_dir", help="批次目录（包含 output/annotated.xlsx）")
    parser.add_argument("--json", action="store_true", help="输出 JSON 摘要")
    args = parser.parse_args()

    batch_dir = args.batch_dir
    xlsx_path = find_batch_output_file(batch_dir)

    if not xlsx_path or not os.path.exists(xlsx_path):
        print(f"❌ 找不到标注文件: {os.path.join(batch_dir, get_batch_output_dir())} 下无 annotated*.xlsx")
        sys.exit(1)

    print(f"🔍 校验中: {xlsx_path}")
    print("=" * 60)

    result = validate_file(xlsx_path)

    if result is None:
        sys.exit(1)

    max_error_rate = get_max_error_rate()
    undecided_threshold = get_undecided_warn_threshold()

    print(f"\n📊 总行数: {result['total_rows']}")
    print(f"❌ 错误行数: {result['error_rows']}")
    print(f"⚠️ 提示行数: {result['warning_rows']}")
    print(f"📈 错误率: {result['error_rate']:.1%}")
    print(f"\n📊 GSB 分布:")
    for k, v in result["gsb_distribution"].items():
        pct = v / result['total_rows'] * 100 if result['total_rows'] > 0 else 0
        print(f"  {k}: {v} 条 ({pct:.1f}%)")

    # 无法区分比例检查
    undecided = result["gsb_distribution"].get("无法区分", 0)
    undecided_pct = undecided / result['total_rows'] if result['total_rows'] > 0 else 0
    if undecided_pct > undecided_threshold:
        print(f"\n⚠️ 警告: '无法区分' 占比 {undecided_pct:.1%} (>{undecided_threshold:.0%})，建议复查是否有应选边的情况")

    if result["errors"]:
        print(f"\n{'='*60}")
        print("详细错误:")
        for row_idx in sorted(result["errors"].keys()):
            print(f"\n--- 行 {row_idx} ---")
            for err in result["errors"][row_idx]:
                print(f"  • {err}")

        # 写入错误文件
        err_path = os.path.join(batch_dir, get_batch_error_file())
        with open(err_path, "w", encoding="utf-8") as f:
            f.write(f"GSB 标注校验报告\n")
            f.write(f"文件: {xlsx_path}\n")
            f.write(f"总行数: {result['total_rows']}, 错误行数: {result['error_rows']}, 错误率: {result['error_rate']:.1%}\n\n")
            for row_idx in sorted(result["errors"].keys()):
                f.write(f"--- 行 {row_idx} ---\n")
                for err in result["errors"][row_idx]:
                    f.write(f"  • {err}\n")
                f.write("\n")
        print(f"\n📝 错误详情已写入: {err_path}")
    else:
        print(f"\n✅ 所有行校验通过！")
        # 清除旧的 errors.txt
        err_path = os.path.join(batch_dir, get_batch_error_file())
        if os.path.exists(err_path):
            os.remove(err_path)

    # 打印启发式提示（不计入错误率，供人工/抽检复核）
    if result["warnings"]:
        print(f"\n{'='*60}")
        print("提示（不阻断流程，建议人工或 quality-audit 复核）:")
        for row_idx in sorted(result["warnings"].keys()):
            print(f"\n--- 行 {row_idx} ---")
            for warn in result["warnings"][row_idx]:
                print(f"  • {warn}")

    # 始终写入 validation-summary.json，供 batch_status / merge 判断校验状态
    summary = {
        "batch_dir": batch_dir,
        "xlsx_path": xlsx_path,
        "total_rows": result["total_rows"],
        "error_rows": result["error_rows"],
        "warning_rows": result["warning_rows"],
        "error_rate": result["error_rate"],
        "gsb_distribution": result["gsb_distribution"],
        "undecided_ratio": undecided_pct,
        "warnings": result["warnings"],
        "passed": result["error_rate"] <= max_error_rate,
    }
    json_path = os.path.join(batch_dir, "validation-summary.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    if args.json:
        print(f"📋 JSON 摘要已写入: {json_path}")

    # 退出码：错误率必须 <= max_error_rate
    if result["error_rate"] > max_error_rate:
        print(f"\n❌ 错误率 {result['error_rate']:.1%} 超过阈值 {max_error_rate:.1%}，必须修正后才能进入下一步")
        sys.exit(1)
    else:
        print(f"\n✅ 错误率 {result['error_rate']:.1%} <= {max_error_rate:.1%}，通过")
        sys.exit(0)


if __name__ == "__main__":
    main()
