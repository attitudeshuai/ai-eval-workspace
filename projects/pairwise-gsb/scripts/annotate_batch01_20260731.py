"""
标注 batch-01（2026-07-31，session 0724）。

基于对生成图的程序化分析（OCR + Vision 分类 + 颜色分区 + 连通区域检测 + 构图分析）
独立完成 GSB 标注，不参考任何历史标注结果。

数据：行2~11（序号1~10），全部为 T2I 任务（无输入图），一致性=不涉及。
"""
import openpyxl
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _config import format_batch_output_file

BATCH_DIR = Path("/Users/lilixian/工作相关/AI/ai-eval-workspace/sessions/pairwise-gsb/0724/2026-07-31/batch-01")
INPUT = BATCH_DIR / "input" / "items_行2-11.xlsx"
OUTPUT = BATCH_DIR / "output" / "annotated_行2-11.xlsx"

# 序号 -> (整体, 指令遵循, 指令归因, 一致性, 一致性归因, 视觉效果, 视觉归因, reason)
ANNOTATIONS = {
    1: dict(
        overall="图片1更好",
        instruction="图片1更好",
        inst_tags="指令响应--其他内容响应差",
        consistency="不涉及",
        cons_tags="",
        visual="无法区分",
        vis_tags="",
        reason="指令遵循：图1包含\u201c夸父炸串\u201d\u201c酸辣粉仅8块9\u201d\u201c微信上下单直接群里下单了可送到学校门口\u201d等文字，内容完整且无错字；图2虽含相同宣传语，但底部出现\u201c炸父\u201d及\u201c预留放置社群号位码\u201d（应为\u201c二维码\u201d）等错字乱码文字，文字内容有误。\n一致性：不涉及。\n视觉效果：图1为橙红色调宣传单，中部为食物特写、上下部为文字横幅，构图清晰；图2同为橙红宣传单、含食物特写与文字，两图均无明显结构或画质问题，画面质量接近。\n综合判断：图2存在明确的错字乱码文字，图1文字完整准确，因此图片1更好。",
    ),
    2: dict(
        overall="图片2更好",
        instruction="无法区分",
        inst_tags="",
        consistency="不涉及",
        cons_tags="",
        visual="图片2更好",
        vis_tags="视觉效果-自然真实-生成效果不自然（假）",
        reason="指令遵循：图1为颁奖大会场景，女孩站在台上，含\u201c三好学生\u201d\u201c100分\u201d等要素；图2为教室黑板前场景，女孩手持\u201c数学试卷100\u201d，黑板上有\u201c三好学生\u201d\u201c荣誉证书\u201d等字样，两图均完成\u201c女孩站在讲台上拿着三好学生奖和满分试卷\u201d的核心要求，场景表达不同但均达标，无法区分。\n一致性：不涉及。\n视觉效果：图1台上横幅文字出现大量乱码（如\u201c期国小宴经典\u201d\u201c按东损王漤米拭\u201d\u201c停錯餐\u201d等不可读文字），画面不自然；图2黑板文字整体可读，仅个别错字（如\u201c好学多间\u201d应为\u201c好问\u201d），画面较干净自然。\n综合判断：图1横幅乱码文字为显著画面瑕疵，图2整体更自然，因此图片2更好。",
    ),
    3: dict(
        overall="图片1更好",
        instruction="无法区分",
        inst_tags="",
        consistency="不涉及",
        cons_tags="",
        visual="图片1更好",
        vis_tags="视觉效果-设计场景中设计感差",
        reason="指令遵循：图1为学生冲刺学习场景，含\u201c冲刺期末，决战复习周！\u201d标语与倒计时数字；图2为复习计划表式海报，含背书计划、倒计时、作息表等内容，两图均围绕\u201c冲刺期末考试\u201d主题，均完成指令，无法区分。\n一致性：不涉及。\n视觉效果：图1为学生与书桌场景，人物、时钟等元素自然，底部仅少量小字；图2文字极为密集、排版拥挤，且出现\u201c时彰政治\u201d\u201c金加以社\u201d\u201c件线\u201d等多处错字乱码，版式设计混乱、可读性差。\n综合判断：图2文字密集且错字乱码多、设计质量差，图1画面更自然清晰，因此图片1更好。",
    ),
    4: dict(
        overall="无法区分",
        instruction="无法区分",
        inst_tags="",
        consistency="不涉及",
        cons_tags="",
        visual="无法区分",
        vis_tags="",
        reason="指令遵循：图1含\u201c早晚刷牙，牙齿亮晶晶\u201d\u201c刷够2分钟哦～\u201d，图2含\u201c早晚刷牙，牙齿亮晶晶！\u201d\u201c每天2次，每次2分钟\u201d，两图文字均完整正确，均完成\u201c提醒孩子刷牙\u201d的要求，无法区分。\n一致性：不涉及。\n视觉效果：图1为牙齿牙刷卡通插画，画面明亮简洁；图2为小孩刷牙卡通形象、画面清晰，两图均无明显结构、画质或自然度问题。\n综合判断：两图核心内容与画面质量相当，无充分依据选边，因此无法区分。",
    ),
    5: dict(
        overall="图片2更好",
        instruction="图片2更好",
        inst_tags="指令响应--物体/图像等具象元素响应差",
        consistency="不涉及",
        cons_tags="",
        visual="无法区分",
        vis_tags="",
        reason="指令遵循：图1为夫子庙/秦淮河夜景画面，以植被、夜空为主，未见明确的南京板鸭主体，核心主体未呈现；图2中央为油亮枣红的板鸭主体，右下标注\u201c南京板鸭\u201d，背景为虚化夜景，完整呈现\u201c南京板鸭\u201d核心要求。\n一致性：不涉及。\n视觉效果：图1为夜景氛围、下部有大面积暗部且主体不明确；图2板鸭主体清晰、油亮质感明显，但两图画面风格不同，整体视觉质量难以明确分出优劣。\n综合判断：图2完成核心主体\u201c南京板鸭\u201d，图1核心主体缺失，因此图片2更好。",
    ),
    6: dict(
        overall="图片1更好",
        instruction="图片1更好",
        inst_tags="指令响应--其他内容响应差",
        consistency="不涉及",
        cons_tags="",
        visual="无法区分",
        vis_tags="",
        reason="指令遵循：图1完整包含\u201c电销精英招聘\u201d\u201c朝九晚七\u201d\u201c大小休\u201d\u201c法定节假日\u201d\u201c无责底薪3k\u201d\u201c平均薪资8-10k\u201d\u201c欢迎各位自荐推荐\u201d等信息，文字正确；图2信息完整且更丰富，但出现\u201c更好的平的+台\u201d（应为\u201c平台\u201d）错字及底部\u201c从迎合似三仔扣\u201d等乱码文字。\n一致性：不涉及。\n视觉效果：图1为蓝色系简洁卡片式海报，层次清晰；图2同为蓝色系、信息密集但版式完整、有卡片分区，两图均无明显结构或画质问题，难以明确分出优劣。\n综合判断：图2存在明确错字与乱码文字，图1文字准确，因此图片1更好。",
    ),
    7: dict(
        overall="图片2更好",
        instruction="图片2更好",
        inst_tags="指令响应--完全未响应",
        consistency="不涉及",
        cons_tags="",
        visual="无法区分",
        vis_tags="",
        reason="指令遵循：图1为夜景图画（含月亮、床铺、诗人等），但画面无任何古诗文字，未完成\u201c抄写古诗\u201d的核心要求；图2含完整《静夜思》全文（\u201c床前明月光，疑是地上霜。举头望明月，低头思故乡。\u201d）并配夜景插图，完成\u201c抄写古诗+图文并茂\u201d指令。\n一致性：不涉及。\n视觉效果：图1画面简洁干净但仅含单一夜景；图2古诗文字与插图结合、版式清晰，两图均无明显结构或画质问题，难以明确分出优劣。\n综合判断：图1缺失古诗文字这一核心内容，因此图片2更好。",
    ),
    8: dict(
        overall="图片1更好",
        instruction="图片1更好",
        inst_tags="指令响应--物体/图像等具象元素响应差",
        consistency="不涉及",
        cons_tags="",
        visual="无法区分",
        vis_tags="",
        reason="指令遵循：图1完整包含安徽顺森环保技术有限公司、中标金额5363206元、项目（安旌高速宁旌段路基工程AJLJ—01标黄砂采购）、招标人及日期2024年11月29日，且右下角有圆形红色公章这一具象元素；图2含公司名、金额、项目、招标人、日期，但右下角未见清晰的圆形红色公章（仅边缘红色条带与零星红点），缺少该具象元素。\n一致性：不涉及。\n视觉效果：图1为白底正式通知书，文字排版规整、条款完整；图2同为白底文书但文字较简略，两图均无明显结构或画质问题，难以明确分出优劣。\n综合判断：图2缺少右下角圆形红色公章且文字较简略，因此图片1更好。",
    ),
    9: dict(
        overall="图片2更好",
        instruction="图片2更好",
        inst_tags="",
        consistency="不涉及",
        cons_tags="",
        visual="图片2更好",
        vis_tags="",
        reason="指令遵循：图1为水彩贺卡，含气球与\u201c节日快乐\u201d字样，色彩较柔和；图2为水彩贺卡，含彩蛋等元素与\u201c节日快乐\u201d字样，色彩饱和度更高、色相对比更鲜明，更贴合\u201c要有纯度明度色相对比\u201d的要求。\n一致性：不涉及。\n视觉效果：图2色彩明快鲜艳、童趣感更强，图1画面简洁柔美，两图均无结构或画质问题，但图2在色彩纯度与对比上更突出。\n综合判断：图2在色彩纯度、明度与色相对比上更符合要求，因此图片2更好。",
    ),
    10: dict(
        overall="图片1更好",
        instruction="图片1更好",
        inst_tags="指令响应--其他内容响应差",
        consistency="不涉及",
        cons_tags="",
        visual="图片1更好",
        vis_tags="",
        reason="指令遵循：图1为三栏排版的封面背景（对应高分冲刺风、简洁大字风、学霸干货风），画面无任何文字且中间预留标题区域，符合\u201c3张且无文字\u201d要求；图2画面出现大量不可读英文乱码文字（如\u201cCheare haree\u201d等），违反\u201c画面无任何文字\u201d的明确要求。\n一致性：不涉及。\n视觉效果：图1三栏背景简洁干净、预留区域明显；图2画面含杂乱乱码文字、不够干净自然，但两图背景主体均无明显结构问题。\n综合判断：图2违反\u201c画面无任何文字\u201d核心要求，因此图片1更好。",
    ),
}


def main():
    wb = openpyxl.load_workbook(INPUT)
    ws = wb.active

    # 定位输出列（按表头）
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            headers[str(cell.value).strip()] = col_idx

    cols = {
        "overall": headers["整体gsb结果"],
        "instruction": headers["分维度gsb/指令遵循"],
        "inst_tags": headers["分维度gsb/指令遵循归因标签"],
        "consistency": headers["分维度gsb/一致性"],
        "cons_tags": headers["分维度gsb/一致性归因标签"],
        "visual": headers["分维度gsb/视觉效果"],
        "vis_tags": headers["分维度gsb/视觉效果归因标签"],
        "reason": headers["reason"],
        "evaluator": headers["评测人"],
    }

    filled = 0
    for row_idx in range(2, ws.max_row + 1):
        seq = ws.cell(row=row_idx, column=1).value
        if seq is None:
            continue
        ann = ANNOTATIONS.get(int(seq))
        if ann is None:
            print(f"⚠️ 行{row_idx} 序号{seq} 无标注，跳过")
            continue
        ws.cell(row=row_idx, column=cols["overall"], value=ann["overall"])
        ws.cell(row=row_idx, column=cols["instruction"], value=ann["instruction"])
        ws.cell(row=row_idx, column=cols["inst_tags"], value=ann["inst_tags"] or None)
        ws.cell(row=row_idx, column=cols["consistency"], value=ann["consistency"])
        ws.cell(row=row_idx, column=cols["cons_tags"], value=ann["cons_tags"] or None)
        ws.cell(row=row_idx, column=cols["visual"], value=ann["visual"])
        ws.cell(row=row_idx, column=cols["vis_tags"], value=ann["vis_tags"] or None)
        ws.cell(row=row_idx, column=cols["reason"], value=ann["reason"])
        ws.cell(row=row_idx, column=cols["evaluator"], value="AI")
        filled += 1
        print(f"✅ 行{row_idx} 序号{seq} 整体={ann['overall']}")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\n已保存: {OUTPUT}（共填充 {filled} 条）")


if __name__ == "__main__":
    main()
