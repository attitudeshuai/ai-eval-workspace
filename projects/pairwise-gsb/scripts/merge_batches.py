"""
合并各批次的标注输出 Excel 为一个完整交付 Excel，并生成 summary.md。

交付文件只保留：deliver_keep_input_columns（默认 序号/prompt_cn/prompt_en）+ output_columns。

用法:
    python merge_batches.py <batches_dir> [--output xxx.xlsx] [--strict]

示例:
    python merge_batches.py deliverables/pairwise-gsb/0724/2026-07-27/
    python merge_batches.py deliverables/pairwise-gsb/0724/2026-07-27/ --strict
"""

import sys
import os
import glob
import json
import argparse
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    sys.exit(1)

from _config import (
    get_batch_output_dir,
    get_difficult_cases_file,
    get_deliver_keep_input_columns,
    get_output_columns,
    get_gsb_values,
    get_tag_lists,
    format_full_output_file,
)


def add_deliverable_validations(ws, headers, last_row):
    """
    给交付文件写入下拉多选数据验证：
    GSB 列 → 图片1更好/图片2更好/无法区分；三个归因标签列 → 对应标签列表。
    保证交付 Excel 中这些列仍是可勾选的 list，而不是纯文字。
    """
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    if last_row < 2:
        return

    tags = get_tag_lists()
    gsb_formula = '"' + ",".join(get_gsb_values()) + '"'
    list_map = {
        "整体gsb结果": gsb_formula,
        "分维度gsb/指令遵循": gsb_formula,
        "分维度gsb/一致性": gsb_formula,
        "分维度gsb/视觉效果": gsb_formula,
        "分维度gsb/指令遵循归因标签": '"' + ",".join(tags["instruction"]) + '"',
        "分维度gsb/一致性归因标签": '"' + ",".join(tags["consistency"]) + '"',
        "分维度gsb/视觉效果归因标签": '"' + ",".join(tags["visual"]) + '"',
    }

    for col_idx, header in enumerate(headers, start=1):
        name = str(header).strip() if header else ""
        if name not in list_map:
            continue
        dv = DataValidation(type="list", formula1=list_map[name], allowBlank=True, showErrorMessage=True)
        ws.add_data_validation(dv)
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")


def find_batch_files(batches_dir):
    """查找所有批次标注输出 Excel，返回 [(batch_name, path), ...]"""
    patterns = [
        os.path.join(batches_dir, "batch-*", "annotated*.xlsx"),
        os.path.join(batches_dir, "batch-*", get_batch_output_dir(), "annotated*.xlsx"),
    ]
    files = []
    for pattern in patterns:
        for f in sorted(glob.glob(pattern)):
            batch_name = os.path.basename(os.path.dirname(f))
            if batch_name.startswith("batch-"):
                files.append((batch_name, f))
    # 去重
    seen = set()
    unique = []
    for name, path in files:
        if name not in seen:
            seen.add(name)
            unique.append((name, path))
    return unique


def check_batch_validation(batch_dir):
    """检查批次是否已通过校验（errors.txt 不存在即视为通过）"""
    errors_path = os.path.join(batch_dir, "errors.txt")
    if not os.path.exists(errors_path):
        return True, "无 errors.txt"
    return False, "errors.txt 存在错误记录"


def count_difficult_cases(batches_dir):
    """统计疑难 case 数量"""
    dc_path = os.path.join(os.path.dirname(batches_dir.rstrip("/\\")), get_difficult_cases_file())
    if not os.path.exists(dc_path):
        return 0
    try:
        wb = openpyxl.load_workbook(dc_path)
        ws = wb.active
        count = max(0, ws.max_row - 1)
        wb.close()
        return count
    except Exception:
        return 0


def get_keep_column_indices(headers):
    """根据表头计算交付保留列的索引（1-based），保持原列顺序。"""
    keep_names = get_deliver_keep_input_columns() + get_output_columns()
    keep_set = set(keep_names)
    return [i for i, h in enumerate(headers, start=1) if h and str(h).strip() in keep_set]


def merge_batches(batches_dir, output_path=None, strict=False):
    batch_files = find_batch_files(batches_dir)

    if not batch_files:
        print(f"❌ 在 {batches_dir} 中未找到 batch-*/annotated*.xlsx")
        sys.exit(1)

    # 严格模式：检查每个批次是否已通过校验
    if strict:
        print("🔒 严格模式：检查各批次校验状态")
        failed = []
        for batch_name, path in batch_files:
            batch_dir = os.path.dirname(path)
            # output 子目录结构时，批次目录是上一级
            if os.path.basename(batch_dir) == get_batch_output_dir():
                batch_dir = os.path.dirname(batch_dir)
            ok, msg = check_batch_validation(batch_dir)
            status = "✅" if ok else "❌"
            print(f"  {status} {batch_name}: {msg}")
            if not ok:
                failed.append(batch_name)
        if failed:
            print(f"\n❌ 以下批次未通过校验，拒绝合并: {failed}")
            sys.exit(1)
        print()

    print(f"合并 {len(batch_files)} 个批次:")
    total_rows = 0
    batch_info = []
    for batch_name, f in batch_files:
        wb = openpyxl.load_workbook(f)
        rows = wb.active.max_row - 1
        wb.close()
        batch_info.append({"name": batch_name, "rows": rows})
        total_rows += rows
        print(f"  {batch_name}: {rows} 条")

    # 序号列（用于统计行号范围）
    merged = openpyxl.Workbook()
    merged_ws = merged.active
    merged_ws.title = "annotated"

    gsb_count = Counter()
    serial_values = []
    keep_cols = None
    out_row = 1

    for i, (batch_name, f) in enumerate(batch_files):
        wb = openpyxl.load_workbook(f)
        ws = wb.active

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if keep_cols is None:
            keep_cols = get_keep_column_indices(headers)
            headers_map = {str(h).strip(): idx for idx, h in enumerate(headers, start=1) if h}
            gsb_col = headers_map.get("整体gsb结果")
            serial_col = headers_map.get("序号")
            # 写表头（仅保留列）
            for new_c, old_c in enumerate(keep_cols, start=1):
                merged_ws.cell(row=1, column=new_c, value=headers[old_c - 1])
            out_row = 2

        for r in range(2, ws.max_row + 1):
            # 跳过空行
            if all(ws.cell(row=r, column=c).value in (None, "") for c in keep_cols):
                continue
            for new_c, old_c in enumerate(keep_cols, start=1):
                merged_ws.cell(row=out_row, column=new_c, value=ws.cell(row=r, column=old_c).value)
            gsb_val = str(ws.cell(row=r, column=gsb_col).value or "") if gsb_col else ""
            if gsb_val:
                gsb_count[gsb_val] += 1
            if serial_col:
                serial_values.append(ws.cell(row=r, column=serial_col).value)
            out_row += 1

        wb.close()

    # 全量文件名：带序号范围
    if output_path is None:
        nums = [int(s) for s in serial_values if isinstance(s, (int, float)) or (isinstance(s, str) and s.isdigit())]
        if nums:
            full_name = format_full_output_file(min(nums), max(nums))
        else:
            full_name = format_full_output_file("?", "?")
        output_path = os.path.join(batches_dir, full_name)

    # 写入下拉多选数据验证（GSB 列 + 三个归因标签列）
    final_headers = [merged_ws.cell(row=1, column=c).value for c in range(1, merged_ws.max_column + 1)]
    add_deliverable_validations(merged_ws, final_headers, out_row - 1)

    merged.save(output_path)
    merged.close()

    # 统计汇总
    print(f"\n{'='*60}")
    print(f"✅ 合并完成: {total_rows} 条 → {output_path}")
    print(f"   交付列: {[merged_ws.cell(row=1, column=c).value for c in range(1, merged_ws.max_column + 1)]}")

    print(f"\nGSB 分布:")
    for k, v in gsb_count.most_common():
        pct = v / total_rows * 100 if total_rows > 0 else 0
        print(f"  {k}: {v} 条 ({pct:.1f}%)")

    # 统计疑难 case
    difficult_count = count_difficult_cases(batches_dir)

    # 生成 summary.md
    summary_path = os.path.join(batches_dir, "summary.md")
    lines = [
        "# GSB 标注交付摘要",
        "",
        f"- **交付目录**: `{batches_dir}`",
        f"- **生成时间**: {datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')}",
        f"- **总条数**: {total_rows}",
        f"- **批次数**: {len(batch_files)}",
        f"- **疑难 case 数**: {difficult_count}",
        "",
        "## 批次明细",
        "",
        "| 批次 | 条数 |",
        "|------|------|",
    ]
    for info in batch_info:
        lines.append(f"| {info['name']} | {info['rows']} |")
    lines.extend([
        "",
        "## GSB 分布",
        "",
        "| 结果 | 条数 | 占比 |",
        "|------|------|------|",
    ])
    for k, v in gsb_count.most_common():
        pct = v / total_rows * 100 if total_rows > 0 else 0
        lines.append(f"| {k} | {v} | {pct:.1f}% |")
    lines.extend([
        "",
        "## 校验状态",
        "",
    ])
    if strict:
        lines.append("- 模式: 严格模式（所有批次必须通过校验）")
    else:
        lines.append("- 模式: 普通模式")
    lines.append("- 详细校验报告见各批次 `errors.txt`（如有）")
    lines.append("")

    with open(summary_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    print(f"\n📋 摘要已生成: {summary_path}")


def main():
    parser = argparse.ArgumentParser(description="合并批次标注结果")
    parser.add_argument("batches_dir", help="包含 batch-*/ 的目录")
    parser.add_argument("--output", "-o", default=None, help="输出文件路径")
    parser.add_argument("--strict", action="store_true", help="严格模式：未通过校验的批次拒绝合并")
    args = parser.parse_args()

    if not os.path.isdir(args.batches_dir):
        print(f"❌ 目录不存在: {args.batches_dir}")
        sys.exit(1)

    merge_batches(args.batches_dir, args.output, args.strict)


if __name__ == "__main__":
    main()
