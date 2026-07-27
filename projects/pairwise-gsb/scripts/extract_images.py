"""
从 Excel 中提取嵌入图片或下载 URL 图片，保存到指定目录。

用法:
    python extract_images.py <excel_path> <output_dir>
    python extract_images.py sessions/pairwise-gsb/0724/2026-07-27/batch-01/input/items.xlsx sessions/pairwise-gsb/0724/2026-07-27/batch-01/images/

输出:
    images/img_0001.png    # 第1张图片（嵌入或下载）
    images/img_0002.png    # 第2张图片
    images/manifest.json   # 图片→行列映射
"""

import sys
import os
import json
import io
import argparse
import re
from pathlib import Path
from urllib.parse import urlparse

sys.path.insert(0, str(Path(__file__).resolve().parent))

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
    print("⚠️ 未安装 Pillow，将按原始格式保存图片")
    print("   建议安装: pip install Pillow")

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
    print("⚠️ 未安装 requests，URL 图片将无法下载")
    print("   建议安装: pip install requests")

from _config import get_image_columns, format_image_filename


def is_url(value):
    """判断单元格值是否为 URL。"""
    if not value:
        return False
    s = str(value).strip()
    return s.startswith("http://") or s.startswith("https://")


def get_extension_from_url(url):
    """从 URL 猜测图片扩展名。"""
    parsed = urlparse(url)
    path = parsed.path.lower()
    for ext in [".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"]:
        if path.endswith(ext):
            return ext
    return ".png"


def get_extension_from_bytes(data):
    """从字节数据猜测图片格式。"""
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if data.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if data.startswith(b"GIF87a") or data.startswith(b"GIF89a"):
        return ".gif"
    if data.startswith(b"RIFF") and data[8:12] == b"WEBP":
        return ".webp"
    return ".png"


def download_image(url, timeout=30):
    """下载 URL 图片，返回 (bytes, content_type) 或 (None, error_msg)。"""
    if not HAS_REQUESTS:
        return None, "requests 未安装"
    try:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        resp = requests.get(url, headers=headers, timeout=timeout, stream=True)
        resp.raise_for_status()
        return resp.content, resp.headers.get("content-type", "")
    except Exception as e:
        return None, str(e)


def extract_images_from_excel(xlsx_path, output_dir, min_row=None, max_row=None, row_offset=0):
    """
    从 Excel 中提取所有嵌入图片和 URL 图片。

    Args:
        xlsx_path: Excel 路径
        output_dir: 输出目录
        min_row/max_row: 仅提取该原始行范围内的图片（含边界，按 Excel 实际行号计）
        row_offset: 写入 manifest 时行号减去该值（用于批次内行号重映射）

    Returns:
        manifest: list[dict], 每个元素 {filename, sheet, row, col, col_letter, col_type, source, width, height, size_bytes}
    """
    wb = openpyxl.load_workbook(xlsx_path)
    os.makedirs(output_dir, exist_ok=True)

    image_columns = get_image_columns()

    def in_range(row):
        if row <= 0:
            return True  # 位置未知的图片不过滤
        if min_row is not None and row < min_row:
            return False
        if max_row is not None and row > max_row:
            return False
        return True

    manifest = []
    img_count = 0
    name_counter = {}  # (row, col_type) -> 已用次数，用于同单元格多图去重

    def next_filename(row, col_type, ext):
        key = (row, str(col_type))
        name_counter[key] = name_counter.get(key, 0) + 1
        return format_image_filename(row, col_type, ext, index=name_counter[key])

    # 先处理嵌入图片
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        images = ws._images if hasattr(ws, "_images") else []

        for img in images:
            # 获取图片位置信息
            try:
                anchor = img.anchor
                if hasattr(anchor, "_from"):
                    row = anchor._from.row + 1
                    col = anchor._from.col + 1
                    col_letter = get_column_letter(col)
                    header = ws.cell(row=1, column=col).value
                    col_type = str(header).strip() if header else f"列{col}"
                else:
                    row, col, col_letter, col_type = 0, 0, "?", "未知"
            except Exception:
                row, col, col_letter, col_type = 0, 0, "?", "未知"

            if not in_range(row):
                continue

            img_count += 1

            # 获取图片尺寸
            width = getattr(img, "width", 0)
            height = getattr(img, "height", 0)

            # 保存图片
            # 文件名使用原始 Excel 行号（保证 row_XXX 与原始文件行号一致），
            # manifest 中 row 为批次内行号、original_row 为原始行号。
            try:
                image_data = img._data()
                ext = get_extension_from_bytes(image_data)
                out_row = row - row_offset if row > 0 else row
                filename = next_filename(row, col_type, ext)
                filepath = os.path.join(output_dir, filename)

                if HAS_PIL:
                    pil_img = Image.open(io.BytesIO(image_data))
                    pil_img.save(filepath, "PNG" if ext == ".png" else "JPEG")
                else:
                    with open(filepath, "wb") as f:
                        f.write(image_data)

                file_size = os.path.getsize(filepath)
                print(f"  ✅ [{img_count:3d}] 原始行{row:3d}(批内行{out_row:3d}) 列{col_letter:2s}({col_type}) → {filename} ({file_size:,} bytes) [嵌入]")

                manifest.append({
                    "filename": filename,
                    "sheet": sheet_name,
                    "row": out_row,
                    "original_row": row,
                    "col": col,
                    "col_letter": col_letter,
                    "col_type": col_type,
                    "source": "embedded",
                    "width": width,
                    "height": height,
                    "size_bytes": file_size,
                })

            except Exception as e:
                print(f"  ❌ [{img_count:3d}] 行{row:3d} 列{col_letter:2s} → 保存失败: {e}")

    # 再处理 URL 图片
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]

        for col_idx, header in enumerate(headers, start=1):
            if header not in image_columns:
                continue

            for row_idx in range(2, ws.max_row + 1):
                if not in_range(row_idx):
                    continue
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if not is_url(cell_value):
                    continue

                img_count += 1
                url = str(cell_value).strip()
                col_letter = get_column_letter(col_idx)

                data, info = download_image(url)
                if data is None:
                    print(f"  ❌ [{img_count:3d}] 行{row_idx:3d} 列{col_letter:2s}({header}) → 下载失败: {info}")
                    continue

                ext = get_extension_from_url(url) or get_extension_from_bytes(data)
                out_row = row_idx - row_offset
                filename = next_filename(row_idx, header, ext)
                filepath = os.path.join(output_dir, filename)

                try:
                    if HAS_PIL:
                        pil_img = Image.open(io.BytesIO(data))
                        pil_img.save(filepath, "PNG" if ext == ".png" else "JPEG")
                    else:
                        with open(filepath, "wb") as f:
                            f.write(data)

                    file_size = os.path.getsize(filepath)
                    width, height = 0, 0
                    if HAS_PIL:
                        try:
                            with Image.open(filepath) as im:
                                width, height = im.size
                        except Exception:
                            pass

                    print(f"  ✅ [{img_count:3d}] 原始行{row_idx:3d}(批内行{out_row:3d}) 列{col_letter:2s}({header}) → {filename} ({file_size:,} bytes) [URL]")

                    manifest.append({
                        "filename": filename,
                        "sheet": sheet_name,
                        "row": out_row,
                        "original_row": row_idx,
                        "col": col_idx,
                        "col_letter": col_letter,
                        "col_type": header,
                        "source": "url",
                        "url": url,
                        "width": width,
                        "height": height,
                        "size_bytes": file_size,
                    })

                except Exception as e:
                    print(f"  ❌ [{img_count:3d}] 行{row_idx:3d} 列{col_letter:2s} → 保存失败: {e}")

    # 写入 manifest
    manifest_path = os.path.join(output_dir, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump({
            "source": os.path.abspath(xlsx_path),
            "total_images": len(manifest),
            "images": manifest,
        }, f, ensure_ascii=False, indent=2)

    wb.close()
    return manifest


def main():
    parser = argparse.ArgumentParser(description="从 Excel 提取嵌入图片或下载 URL 图片")
    parser.add_argument("excel_path", help="Excel 文件路径")
    parser.add_argument("output_dir", help="图片输出目录")
    parser.add_argument("--min-row", type=int, default=None, help="仅提取该原始行及之后的图片")
    parser.add_argument("--max-row", type=int, default=None, help="仅提取该原始行及之前的图片")
    parser.add_argument("--row-offset", type=int, default=0, help="manifest 中行号偏移（批次内重映射用）")
    args = parser.parse_args()

    xlsx_path = args.excel_path
    output_dir = args.output_dir

    if not os.path.exists(xlsx_path):
        print(f"❌ Excel 文件不存在: {xlsx_path}")
        sys.exit(1)

    print(f"📂 源文件: {xlsx_path}")
    print(f"📁 输出目录: {output_dir}")
    if args.min_row or args.max_row or args.row_offset:
        print(f"🔢 行范围: {args.min_row or '-'} ~ {args.max_row or '-'}, 偏移: {args.row_offset}")
    print(f"{'='*60}")

    manifest = extract_images_from_excel(
        xlsx_path, output_dir,
        min_row=args.min_row, max_row=args.max_row, row_offset=args.row_offset,
    )

    print(f"{'='*60}")
    print(f"✅ 提取完成: {len(manifest)} 张图片 → {output_dir}")
    print(f"📋 映射清单: {os.path.join(output_dir, 'manifest.json')}")

    # 按列统计
    from collections import Counter
    cols = Counter(img["col_letter"] for img in manifest)
    sources = Counter(img["source"] for img in manifest)
    print(f"\n📊 按列分布:")
    for col, count in sorted(cols.items()):
        print(f"  列 {col}: {count} 张")
    print(f"\n📊 按来源分布:")
    for src, count in sources.items():
        print(f"  {src}: {count} 张")


if __name__ == "__main__":
    main()
