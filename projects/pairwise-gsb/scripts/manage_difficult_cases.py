"""
疑难 case 管理脚本。

用法:
    python manage_difficult_cases.py add <batch_dir> <row_in_batch> <reason>
    python manage_difficult_cases.py resolve <difficult_cases.xlsx> <row_id> <decision>
    python manage_difficult_cases.py list <session_date_dir>

示例:
    python manage_difficult_cases.py add sessions/pairwise-gsb/0724/2026-07-27/batch-01 3 "两图均有多手指，无法判断优劣"
    python manage_difficult_cases.py resolve sessions/pairwise-gsb/0724/2026-07-27/difficult-cases.xlsx 3 "两图问题相近，判无法区分"
    python manage_difficult_cases.py list sessions/pairwise-gsb/0724/2026-07-27
"""

import sys
import os
import json
import argparse
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

try:
    import openpyxl
except ImportError:
    print("需要 openpyxl: pip install openpyxl")
    sys.exit(1)

from _config import (
    get_difficult_cases_file,
    get_difficult_cases_log,
    find_batch_input_file,
    find_batch_output_file,
)


def get_session_dir_from_batch(batch_dir):
    """从批次目录推断 session/date 目录"""
    batch_path = Path(batch_dir).resolve()
    return batch_path.parent


def init_difficult_cases_file(path):
    """初始化疑难 case Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "difficult_cases"
    headers = [
        "case_id", "session_date", "batch", "row_in_batch", "row_in_original",
        "prompt", "reason", "status", "decision", "created_at", "resolved_at",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    wb.save(path)
    wb.close()


def load_or_create_difficult_cases(session_dir):
    """加载或创建疑难 case 文件"""
    dc_path = Path(session_dir) / get_difficult_cases_file()
    if not dc_path.exists():
        init_difficult_cases_file(dc_path)
    return dc_path


def append_log(session_dir, message):
    """追加疑难 case 处理日志"""
    log_path = Path(session_dir) / get_difficult_cases_log()
    timestamp = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"- [{timestamp}] {message}\n")


def get_original_row_from_batch(batch_dir, row_in_batch):
    """从批次输入 Excel 中读取指定行的原始数据"""
    input_file = find_batch_input_file(batch_dir)
    if not input_file:
        return None
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    row_data = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        value = ws.cell(row=row_in_batch, column=col).value
        row_data[str(header)] = value
    wb.close()
    return row_data


def cmd_add(args):
    """添加疑难 case"""
    batch_dir = Path(args.batch_dir).resolve()
    row_in_batch = int(args.row_in_batch)
    reason = args.reason

    session_dir = get_session_dir_from_batch(batch_dir)
    dc_path = load_or_create_difficult_cases(session_dir)

    batch_name = batch_dir.name
    row_data = get_original_row_from_batch(batch_dir, row_in_batch) or {}
    prompt = str(row_data.get("prompt_cn") or row_data.get("prompt_en") or "")

    wb = openpyxl.load_workbook(dc_path)
    ws = wb.active
    new_row = ws.max_row + 1
    case_id = new_row - 1  # 第一条为 1

    ws.cell(row=new_row, column=1, value=case_id)
    ws.cell(row=new_row, column=2, value=session_dir.name)
    ws.cell(row=new_row, column=3, value=batch_name)
    ws.cell(row=new_row, column=4, value=row_in_batch)
    ws.cell(row=new_row, column=5, value=row_in_batch)  # 简化：暂不映射原始行号
    ws.cell(row=new_row, column=6, value=prompt[:500])
    ws.cell(row=new_row, column=7, value=reason)
    ws.cell(row=new_row, column=8, value="pending")
    ws.cell(row=new_row, column=9, value="")
    ws.cell(row=new_row, column=10, value=datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"))
    ws.cell(row=new_row, column=11, value="")

    wb.save(dc_path)
    wb.close()

    append_log(session_dir, f"新增疑难 case #{case_id}: {batch_name} 行{row_in_batch} - {reason}")
    print(f"✅ 已添加疑难 case #{case_id}")
    print(f"   批次: {batch_name}, 行: {row_in_batch}")
    print(f"   原因: {reason}")
    print(f"   文件: {dc_path}")


def cmd_resolve(args):
    """解决疑难 case 并回填"""
    dc_path = Path(args.difficult_cases).resolve()
    row_id = int(args.row_id)
    decision = args.decision

    if not dc_path.exists():
        print(f"❌ 疑难 case 文件不存在: {dc_path}")
        sys.exit(1)

    session_dir = dc_path.parent
    wb = openpyxl.load_workbook(dc_path)
    ws = wb.active

    # 查找 case
    target_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == row_id:
            target_row = r
            break

    if target_row is None:
        print(f"❌ 未找到 case_id={row_id}")
        sys.exit(1)

    batch_name = ws.cell(row=target_row, column=3).value
    row_in_batch = ws.cell(row=target_row, column=4).value
    status = ws.cell(row=target_row, column=8).value

    if status == "resolved":
        print(f"⚠️ case #{row_id} 已解决，跳过")
        return

    # 更新疑难 case 状态
    ws.cell(row=target_row, column=8, value="resolved")
    ws.cell(row=target_row, column=9, value=decision)
    ws.cell(row=target_row, column=11, value=datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"))
    wb.save(dc_path)
    wb.close()

    # 回填到批次标注输出文件
    batch_dir = session_dir / batch_name
    annotated_file = find_batch_output_file(batch_dir)

    if annotated_file:
        wb2 = openpyxl.load_workbook(annotated_file)
        ws2 = wb2.active
        # 在 reason 列后追加统一口径结论
        headers = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
        reason_col = None
        for i, h in enumerate(headers, start=1):
            if h == "reason":
                reason_col = i
                break
        if reason_col:
            old_reason = ws2.cell(row=row_in_batch, column=reason_col).value or ""
            new_reason = f"{old_reason}\n[疑难 case 统一口径] {decision}".strip()
            ws2.cell(row=row_in_batch, column=reason_col, value=new_reason)
        wb2.save(annotated_file)
        wb2.close()
        print(f"✅ 已回填到 {batch_name} 行 {row_in_batch}")
    else:
        print(f"⚠️ 批次输出文件不存在，仅记录结论: {annotated_file}")

    append_log(session_dir, f"解决疑难 case #{row_id}: {batch_name} 行{row_in_batch} - {decision}")
    print(f"✅ 已解决疑难 case #{row_id}")
    print(f"   批次: {batch_name}, 行: {row_in_batch}")
    print(f"   结论: {decision}")


def cmd_list(args):
    """列出未解决疑难 case"""
    session_dir = Path(args.session_dir).resolve()
    dc_path = session_dir / get_difficult_cases_file()

    if not dc_path.exists():
        print(f"📭 无疑难 case 文件: {dc_path}")
        return

    wb = openpyxl.load_workbook(dc_path)
    ws = wb.active

    pending = []
    for r in range(2, ws.max_row + 1):
        status = ws.cell(row=r, column=8).value
        if status == "pending":
            pending.append({
                "case_id": ws.cell(row=r, column=1).value,
                "batch": ws.cell(row=r, column=3).value,
                "row": ws.cell(row=r, column=4).value,
                "reason": ws.cell(row=r, column=7).value,
                "created_at": ws.cell(row=r, column=10).value,
            })

    wb.close()

    if not pending:
        print(f"✅ 所有疑难 case 已解决")
        return

    print(f"📋 未解决疑难 case ({len(pending)} 条):")
    print(f"{'='*70}")
    for p in pending:
        print(f"  #{p['case_id']} [{p['batch']} 行{p['row']}] {p['reason']}")
        print(f"      创建时间: {p['created_at']}")
    print(f"{'='*70}")
    print(f"文件: {dc_path}")


def main():
    parser = argparse.ArgumentParser(description="疑难 case 管理")
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    p_add = subparsers.add_parser("add", help="添加疑难 case")
    p_add.add_argument("batch_dir", help="批次目录")
    p_add.add_argument("row_in_batch", help="批次内行号")
    p_add.add_argument("reason", help="疑难原因")

    p_resolve = subparsers.add_parser("resolve", help="解决疑难 case")
    p_resolve.add_argument("difficult_cases", help="difficult-cases.xlsx 路径")
    p_resolve.add_argument("row_id", help="case_id")
    p_resolve.add_argument("decision", help="统一口径结论")

    p_list = subparsers.add_parser("list", help="列出未解决疑难 case")
    p_list.add_argument("session_dir", help="session/date 目录")

    args = parser.parse_args()

    if args.command == "add":
        cmd_add(args)
    elif args.command == "resolve":
        cmd_resolve(args)
    elif args.command == "list":
        cmd_list(args)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
